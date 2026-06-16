from openpyxl import load_workbook

from app.models.participant import Participant
from app.models.role_rate import RoleRate
from app.services.excel_import_service import (
    PARTICIPANT_IMPORT_HEADERS,
    ROLE_RATE_IMPORT_HEADERS,
)
from app.services.master_excel_export_service import (
    export_participants_to_excel,
    export_role_rates_to_excel,
)


def test_export_participants_to_excel_writes_template_columns(tmp_path):
    output_path = tmp_path / "participants.xlsx"
    participants = [
        Participant(
            participant_id="P-000001",
            is_active=True,
            name="Yamada Taro",
            department="Sales",
            position="Manager",
            display_name="A",
            hourly_rate=6000,
            sort_order=1,
        ),
        Participant(
            participant_id="P-000002",
            is_active=False,
            name="Sato Hanako",
            hourly_rate=4000,
        ),
    ]

    export_participants_to_excel(participants, output_path)

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    assert rows == [
        PARTICIPANT_IMPORT_HEADERS,
        ("有効", "Yamada Taro", "Sales", "Manager", "A", 6000, 1),
        ("無効", "Sato Hanako", None, None, None, 4000, None),
    ]


def test_export_role_rates_to_excel_writes_template_columns(tmp_path):
    output_path = tmp_path / "role_rates.xlsx"
    role_rates = [
        RoleRate(
            role_rate_id="R-000001",
            is_active=True,
            role_name="Manager",
            hourly_rate=6000,
            sort_order=1,
        ),
        RoleRate(
            role_rate_id="R-000002",
            is_active=False,
            role_name="Staff",
            hourly_rate=3000,
        ),
    ]

    export_role_rates_to_excel(role_rates, output_path)

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    assert rows == [
        ROLE_RATE_IMPORT_HEADERS,
        ("有効", "Manager", 6000, 1),
        ("無効", "Staff", 3000, None),
    ]
