from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from app.models.participant import Participant
from app.models.role_rate import RoleRate
from app.services.duplicate_service import (
    find_participant_duplicate_confirmation_groups,
    find_participant_duplicate_error_groups,
    find_role_rate_duplicate_groups,
)
from app.services.license_settings_service import DEVICE_ROLE_MASTER, validate_device_role


PARTICIPANT_IMPORT_HEADERS = (
    "有効",
    "氏名",
    "部署",
    "役職",
    "識別名",
    "概算時間単価",
    "表示順",
)
ROLE_RATE_IMPORT_HEADERS = (
    "有効",
    "役職名",
    "概算時間単価",
    "表示順",
)

ACTIVE_LABEL = "有効"
INACTIVE_LABEL = "無効"


class ExcelImportError(Exception):
    pass


class ExcelImportPermissionError(ExcelImportError):
    pass


class ExcelImportValidationError(ExcelImportError):
    pass


def import_participants_from_excel(
    file_path: Path,
    device_role: str,
    start_no: int = 0,
) -> list[Participant]:
    _require_master_device(device_role)
    rows = _read_template_rows(file_path, PARTICIPANT_IMPORT_HEADERS)

    participants = []
    for index, row in enumerate(rows, start=1):
        participants.append(
            Participant(
                participant_id=_participant_id(start_no + index),
                is_active=_parse_active(row["有効"], "有効"),
                name=_required_text(row["氏名"], "氏名"),
                department=_optional_text(row["部署"]),
                position=_optional_text(row["役職"]),
                display_name=_optional_text(row["識別名"]),
                hourly_rate=_positive_int(row["概算時間単価"], "概算時間単価"),
                sort_order=_optional_positive_int(row["表示順"], "表示順"),
            )
        )

    _validate_participant_duplicates(participants)
    return participants


def import_role_rates_from_excel(
    file_path: Path,
    device_role: str,
    start_no: int = 0,
) -> list[RoleRate]:
    _require_master_device(device_role)
    rows = _read_template_rows(file_path, ROLE_RATE_IMPORT_HEADERS)

    role_rates = []
    for index, row in enumerate(rows, start=1):
        role_rates.append(
            RoleRate(
                role_rate_id=_role_rate_id(start_no + index),
                is_active=_parse_active(row["有効"], "有効"),
                role_name=_required_text(row["役職名"], "役職名"),
                hourly_rate=_positive_int(row["概算時間単価"], "概算時間単価"),
                sort_order=_optional_positive_int(row["表示順"], "表示順"),
            )
        )

    _validate_role_rate_duplicates(role_rates)
    return role_rates


def _require_master_device(device_role: str) -> None:
    validated_device_role = validate_device_role(device_role)
    if validated_device_role != DEVICE_ROLE_MASTER:
        raise ExcelImportPermissionError("excel import is allowed only on master device")


def _read_template_rows(
    file_path: Path,
    expected_headers: tuple[str, ...],
) -> list[dict[str, Any]]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        header_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            None,
        )
        if header_row is None:
            raise ExcelImportValidationError("template header row is missing")

        headers = tuple(_optional_text(value) for value in header_row)
        if headers != expected_headers:
            raise ExcelImportValidationError("template headers are invalid")

        rows = []
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            if _is_blank_row(values):
                continue
            rows.append(dict(zip(expected_headers, values, strict=True)))
        return rows
    finally:
        workbook.close()


def _validate_participant_duplicates(participants: list[Participant]) -> None:
    if find_participant_duplicate_error_groups(participants):
        raise ExcelImportValidationError("participant identity is duplicated")

    for group in find_participant_duplicate_confirmation_groups(participants):
        if any(not item.display_name.strip() for item in group.items):
            raise ExcelImportValidationError(
                "display_name is required for participants with same name, department, and position"
            )


def _validate_role_rate_duplicates(role_rates: list[RoleRate]) -> None:
    if find_role_rate_duplicate_groups(role_rates):
        raise ExcelImportValidationError("role_name is duplicated")


def _parse_active(value: Any, field_name: str) -> bool:
    text = _required_text(value, field_name)
    if text == ACTIVE_LABEL:
        return True
    if text == INACTIVE_LABEL:
        return False
    raise ExcelImportValidationError(f"{field_name} must be 有効 or 無効")


def _required_text(value: Any, field_name: str) -> str:
    text = _optional_text(value)
    if not text:
        raise ExcelImportValidationError(f"{field_name} is required")
    return text


def _optional_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _positive_int(value: Any, field_name: str) -> int:
    parsed_value = _int_value(value, field_name)
    if parsed_value < 1:
        raise ExcelImportValidationError(f"{field_name} must be at least 1")
    return parsed_value


def _optional_positive_int(value: Any, field_name: str) -> int | None:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    return _positive_int(value, field_name)


def _int_value(value: Any, field_name: str) -> int:
    if isinstance(value, bool):
        raise ExcelImportValidationError(f"{field_name} must be an integer")
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        text = value.strip()
        if text.isdecimal():
            return int(text)
    raise ExcelImportValidationError(f"{field_name} must be an integer")


def _is_blank_row(values: Iterable[Any]) -> bool:
    return all(_optional_text(value) == "" for value in values)


def _participant_id(no: int) -> str:
    return f"P-{no:06d}"


def _role_rate_id(no: int) -> str:
    return f"R-{no:06d}"
